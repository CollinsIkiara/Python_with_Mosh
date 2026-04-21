# This is a project that runs a Python script to automate the process of correcting prices in an Excel workbook and creating a bar chart based on the corrected prices. The script uses the openpyxl library to manipulate the Excel file.

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename) # Load the workbook from the specified filename
    sheet = wb['Sheet1'] # Access the sheet named 'Sheet1' in the workbook


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3) # Access the cell in the 3rd column (price) for the current row
        corrected_price = cell.value * 0.9 # Corrected price: applies a 10% discount to price
        corrected_price_cell = sheet.cell(row, 4) # Access the cell in the 4th column (corrected price) for the current row
        corrected_price_cell.value = corrected_price # Set the value of the corrected price cell to the calculated corrected price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row, 
            min_col=4, 
            max_col=4) # Create a reference to the range of corrected prices in the 4th column, starting from row 2 to the last row with data

    chart = BarChart() # Create a new bar chart object
    chart.add_data(values) # Add the data from the reference to the bar chart
    sheet.add_chart(chart, 'e2') # Add the bar chart to the sheet at the position 'e2'
        
    wb.save(filename) # Save the workbook with the changes made (corrected prices and added chart)
    
process_workbook('transactions2.xlsx') # Call the function to process the workbook named 'transactions2.xlsx' which contains the data to be corrected and visualized.